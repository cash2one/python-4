
# coding=UTF-8
import os;
import copy;
from openpyxl.styles import Color,colors,Fill,PatternFill,Font,fills;
from openpyxl import load_workbook;
from openpyxl import utils
from openpyxl.cell import Cell;
from openpyxl.worksheet.worksheet import Worksheet;
from openpyxl.worksheet import *;
import fileutil;


#这个是这类的函数测试用  因为如果在雷利加代码测试不方便 容易弄错  所以
def read_excel2007(fileName):
    wb = load_workbook(fileName,data_only=True);
    # print("Worksheet range(s):" , wb.get_named_ranges());
    # print("Worksheet name(s):" , wb.get_sheet_names());
    # sheetnames = wb.get_sheet_names();
    ws =  wb.get_active_sheet();

    # print(ws.get_highest_row());
    print(ws.cell(row=1,column=2).value);
    ws.cell(row=3,column=2).value = "jiayou";
    wb.save(filename);
    # print(ws.merged_cell_ranges);
    # print(ws.merged_cells);
    # print(ws.cell('C2').value);
    # print(ws.merged_cell_ranges[1]);
    # print(utils.range_boundaries(ws.merged_cell_ranges[1]));
    # print(utils.column_index_from_string(ws.merged_cells[1]));
    # print(ws.name);
    # print("*****************");



class SheetIsNotExist(Exception):
     def __init__(self, value):
         self.value = value
     def __str__(self):
        return repr(self.value)

class ColumnIsNotVaild(Exception):
     def __init__(self, value):
         self.value = value
     def __str__(self):
        return repr(self.value)

class ReadExcel2007:
    wb = None;
    path = "";
    sheet = None;
    currentSheetName = None;
    mergeCellsColumnAndRow = [];#是这种形式的(2,1,3,2)
    mergeCellsCharAndIndex = [];#是这种形式的(B1:C2)
    currentSheetIndex = 0;#-1代表没有
    def __init__(self,path,sheetIndex = 0):
        if(os.path.exists(path)==False):
            raise fileutil.FileIsNotExist(path + " 不存在 ");

        """
        ***  下面这个data_only是用来指示到底是读取数值 还是公式  但是如果设置为读取数值
        ***  但是原来是公式  那么如果只是读文件  那么没有问题 如果也保存了数据  那么原来的公式 就会被数值替代 没有了公式
        """
        self.wb = load_workbook(path,data_only=True);
        sheetName = self.wb.get_sheet_names();
        # print(sheetName);
        if len(sheetName) != 0:
            self.path = path;
            self.sheet = self.wb.get_sheet_by_name(sheetName[0]);
            self.curentSheetIndex = sheetIndex;
            self.currentSheetName = sheetName[sheetIndex];
            self.__covertStringToColumnAndRow();
            self.__setMergeCellsCharAndIndex();

        else:
            raise SheetIsNotExist("第 "+sheetIndex+" "+"sheet 不存在");


    def setCurrentSheetByIndex(self,sheetIndex):
        sheetName = self.wb.get_sheet_names();
        if sheetIndex >  len(sheetName) and sheetIndex <0:
            raise SheetIsNotExist("第 "+sheetIndex+" "+"sheet 不存在");
        if sheetIndex != self.curentSheetIndex:
            self.curentSheetIndex = sheetIndex;
            self.currentSheetName = sheetName[sheetIndex];
            self.sheet = self.wb.get_sheet_by_name(sheetName[0]);
            self.__covertStringToColumnAndRow();
            self.__setMergeCellsCharAndIndex();

    def setCurrentSheetByName(self,sheetName):
        sheetNameList = self.wb.get_sheet_names();
        if sheetName not in  sheetNameList:
            raise SheetIsNotExist(" 名为 "+sheetName+" "+" 的sheet 不存在");
        if str(sheetName)!=self.currentSheetName:
            suoyin = list(sheetNameList).index(str(sheetName).strip());
            self.curentSheetIndex = suoyin;
            self.currentSheetName = sheetName[sheetName];
            self.sheet = self.wb.get_sheet_by_name(sheetName[0]);
            self.__covertStringToColumnAndRow();
            self.__setMergeCellsCharAndIndex();


    #这个函数会把合并单元格的内容 给每个合并单元格 hang与lie 都是从1 开始
    def getCellValueByColumnAndRow(self,hang,lie):
        for i in self.mergeCellsColumnAndRow:
            if hang>=i[1] and hang<=i[3] and lie>=i[0] and lie <=i[2]:
                return self.sheet.cell(row=i[1],column=i[0]).value;

        # print("else");
        return self.sheet.cell(row=hang,column=lie).value;

    #使用的时候excel 必须关闭 否则会报权限不足 hang与lie 都是从1 开始
    def setCellValueByColumnAndRow(self,hang,lie,value = None):
        for i in self.mergeCellsColumnAndRow:
            if hang>=i[1] and hang<=i[3] and lie>=i[0] and lie <=i[2]:
                self.__setDanYuanGeWidth(utils.get_column_letter(i[0]));
                self.sheet.cell(row=i[1],column=i[0]).value = value;
                self.wb.save(self.path);

        self.__setDanYuanGeWidth(utils.get_column_letter(lie));
        self.sheet.cell(row=hang,column=lie).value = value;
        self.wb.save(self.path);



    #B2这种形式给出
    def getCellValueByCharAndRow(self,stringColumn,row):
        stringColumn = str(stringColumn).upper().strip();
        for i in self.mergeCellsCharAndIndex:
            if stringColumn>=i[0] and stringColumn<=i[2] and row>=i[1] and row <=i[3]:
                return self.sheet.cell(str(i[0]+str(i[1]))).value;

        return self.sheet.cell(str(stringColumn)+str(row)).value;

    #使用的时候excel 必须关闭 否则会报权限不足
    def setCellValueByCharAndRow(self,stringColumn,row,value=None):
        stringColumn = str(stringColumn).upper().strip();
        for i in self.mergeCellsCharAndIndex:
            if stringColumn>=i[0] and stringColumn<=i[2] and row>=i[1] and row <=i[3]:
                pinJieStr = str(i[0]).strip()+str(i[1]).strip();
                self.__setDanYuanGeWidth(str(i[0]).strip());
                self.sheet.cell(pinJieStr).value = value;
                self.wb.save(self.path);
                return;

        self.__setDanYuanGeWidth(str(stringColumn));
        self.sheet.cell(str(stringColumn)+str(row)).value = value;
        self.wb.save(self.path);


    def __setDanYuanGeWidth(self,cellName,widthValue = None):
        cellName.strip();
        self.sheet.cell(coordinate = cellName+str(1));

        # print(self.sheet.column_dimensions);
        width = self.sheet.column_dimensions[cellName].width;
        if(widthValue is not None):
            self.sheet.column_dimensions[cellName].width = widthValue;# self.sheet.column_dimensions[cellName].width;
        elif(width is not None and width != 0):
            self.sheet.column_dimensions[cellName].width = self.sheet.column_dimensions[cellName].width;
        else:
            self.sheet.column_dimensions[cellName].width = 10;

        # self.sheet.column_dimensions[cellName].width = 10;

    #通过“2 3” 2代表第二行 3代表 第三列 这种形式来设置单元格的颜色
    def setCellColorByColumnAndRow(self,hang,lie,R=255,G=255,B=255):
        if(lie<0 ):
            raise ColumnIsNotVaild("列 不能小于0");
        if(hang<0 ):
            raise ColumnIsNotVaild(str(hang)+"不能小于0行");

        if(R==255 and G==255 and B == 255):
            fill = PatternFill(patternType = fills.FILL_NONE,fgColor=Color(rgb=self.rgbConvertToHex(R,G,B)));
        else:
            fill = PatternFill(patternType = fills.FILL_SOLID,fgColor=Color(rgb=self.rgbConvertToHex(R,G,B)));

        for i in self.mergeCellsColumnAndRow:
            if hang>=i[1] and hang<=i[3] and lie>=i[0] and lie <=i[2]:
                for neiHang in range(i[1],i[3]+1):
                    for neiCol in range(i[0],i[2]+1):
                      self.sheet.cell(row=neiHang,column=neiCol).fill = fill;
                self.wb.save(self.path);
                return;

        self.sheet.cell(row=hang,column=lie).fill = fill;
        self.wb.save(self.path);

    #得到 当前单元格的 颜色 有两种可能  就是如果 是纯颜色 那么返回(R，G，B)  如果是主题颜色
    #会返回{'theme':fgColor.value,'persentage':fgColor.tint};
    #其中theme 是主题  就是对应excel 中主体颜色  tint 代表是主题颜色的稀释率  至于主题颜色的RGB值 python 获得不了
    def getCellColorByColumnAndRow(self,hang,lie):
        if(lie<0 ):
            raise ColumnIsNotVaild("列 不能小于0");
        if(hang<0 ):
            raise ColumnIsNotVaild(str(hang)+"不能小于0行");

        cell = self.sheet.cell(row=hang,column=lie);
        for i in self.mergeCellsColumnAndRow:
            if hang>=i[1] and hang<=i[3] and lie>=i[0] and lie <=i[2]:
                cell = self.sheet.cell(row=i[1],column=i[0]);
                if(cell.fill.fgColor.type =='rgb'):
                    return self.__colorHexStringToRGB(cell.fill.fgColor.rgb);
                elif(cell.fill.fgColor.type =='theme'):
                    fgColor = cell.fill.fgColor;
                    return {'theme':fgColor.value,'persentage':fgColor.tint};

        if(cell.fill.fgColor.type =='rgb'):
            return self.__colorHexStringToRGB(cell.fill.fgColor.rgb);
        elif(cell.fill.fgColor.type =='theme'):
            fgColor = cell.fill.fgColor;
            return {'theme':fgColor.value,'persentage':fgColor.tint};

    #将RGB 转换为 #16进制  如 (R，G,B)=(255,255,255) ->"00FFFFFF"
    def rgbConvertToHex(self,R,G,B):
        myHex = "00";#表示alpha 透明度
        myHex+= "%02x" %(R);
        myHex+= "%02x" %(G);
        myHex+= "%02x" %(B);
        return str(myHex).upper();


    #通过“B2”这种形式来设置单元格的颜色
    def setCellColorByCharAndRow(self,stringColumn,row,R=255,G=255,B=255):
        if(utils.column_index_from_string(stringColumn.strip())<0 ):
            raise ColumnIsNotVaild("列 不能小于0");
        if(row<0 ):
            raise ColumnIsNotVaild(str(row)+"不能小于0行");

        if(R==255 and G==255 and B == 255):
            fill = PatternFill(patternType = fills.FILL_NONE,fgColor=Color(rgb=self.rgbConvertToHex(R,G,B)));
        else:
            fill = PatternFill(patternType = fills.FILL_SOLID,fgColor=Color(rgb=self.rgbConvertToHex(R,G,B)));
        stringColumn = str(stringColumn).upper().strip();
        # print("hdsjafaksldf");
        # print(self.mergeCellsColumnAndRow);
        for i in self.mergeCellsCharAndIndex:
            # print("stringcolumn  "+str(stringColumn));
            if stringColumn>=str(i[0]).upper() and stringColumn<=str(i[2]).upper() and row>=i[1] and row <=i[3]:
                for neiHang in range((i[1]),i[3]+1):
                    mybegin = utils.column_index_from_string(i[0]);
                    myend = utils.column_index_from_string(i[2]);
                    for neiCol in range(mybegin,myend+1):
                        self.sheet.cell(row=neiHang,column=neiCol).fill = fill;
                self.wb.save(self.path);
                return;

        self.sheet.cell(str(stringColumn)+str(row)).fill = fill;
        self.wb.save(self.path);

    #得到 当前单元格的 颜色 有两种可能  就是如果 是纯颜色 那么返回(R，G，B)  如果是主题颜色
    # 会返回{'theme':fgColor.value,'persentage':fgColor.tint};
    #  其中theme 是主题  就是对应excel 中主体颜色  tint 代表是主题颜色的稀释率  至于主题颜色的RGB值 python 获得不了
    def getCellColorByCharAndRow(self,stringColumn,row):
        if(utils.column_index_from_string(stringColumn.strip())<0 ):
            raise ColumnIsNotVaild("列 不能小于0");
        if(row<0 ):
            raise ColumnIsNotVaild(str(row)+"不能小于0行");
        stringColumn = str(stringColumn).upper().strip();
        cell = self.sheet.cell(str(stringColumn)+str(row));
        for i in self.mergeCellsCharAndIndex:
            # print("stringcolumn  "+str(stringColumn));
            if stringColumn>=str(i[0]).upper() and stringColumn<=str(i[2]).upper() and row>=i[1] and row <=i[3]:
                cell = self.sheet.cell(str(i[0])+str(i[1]));
                if(cell.fill.fgColor.type =='rgb'):
                    return self.__colorHexStringToRGB(cell.fill.fgColor.rgb) ;
                elif(cell.fill.fgColor.type =='theme'):
                    fgColor = cell.fill.fgColor;
                    return {'theme':fgColor.value,'persentage':fgColor.tint};
        # for i in self.sheet.cell(str(stringColumn)+str(row)).fill.fgColor:
        #     print (i);
        # print(self.sheet.cell(str(stringColumn)+str(row)).fill.fgColor.tint);


        if(cell.fill.fgColor.type=='rgb'):
            return self.__colorHexStringToRGB(cell.fill.fgColor.rgb);
        elif(cell.fill.fgColor.type=='theme'):
            fgColor = cell.fill.fgColor;
            return {'theme':fgColor.value,'persentage':fgColor.tint};

    #用于转换"00ff0000"  到 r = 255 G = 0 B =0;
    def __colorHexStringToRGB(self,hexString):
        if(hexString is None):
             return None;
        print(hexString)
        if(len(hexString)==6):
            R = hexString[0:2];
            G = hexString[2:4];
            B = hexString[4:6];
            R = int(R,16);
            G = int(G,16);
            B = int(B,16);
            return (R,G,B);
        if (len(hexString)==8):
            R = hexString[2:4];
            G = hexString[4:6];
            B = hexString[6:8];
            R = int(R,16);
            G = int(G,16);
            B = int(B,16);
            return (R,G,B);



    #返回当前sheet的整个内容 以列表里面再列表返回  第一层的列表的每个元素对应excel里的列 而不是行
    def getAllSheet(self):
        rowLength = self.sheet.get_highest_row();
        colLength = self.sheet.get_highest_column();
        sheetListHang  = [];
        sheetLisLie  = [];
        # print("colLength"+str(colLength));
        # print("rowLength"+str(rowLength));
        for i in range(1,colLength+1):
            for j in range(1,rowLength+1):
                # print("i= "+str(i)+" j=  "+str(j));
                sheetLisLie.append(self.getCellValueByColumnAndRow(j,i));

            sheetListHang.append(copy.deepcopy(sheetLisLie));
            sheetLisLie.clear();
        return sheetListHang;


    #返回当前sheet的指定行的内容 如果当前行超过了 最大行 或者小于0则会抛出异常
    def getOneRow(self,whichRow):
        if(whichRow<0 ):
            raise ColumnIsNotVaild(str(whichRow)+"不能小于0行");
        if( whichRow >self.sheet.get_highest_column()):
            raise ColumnIsNotVaild(str(whichRow)+"行已经大于"+str(self.sheet.get_highest_row())+"这是最大行");
        colLength = self.sheet.get_highest_column();
        sheetListHang  = [];
        # print("colLength"+str(colLength));
        for i in range(1,colLength+1):
            sheetListHang.append(self.getCellValueByColumnAndRow(whichRow,i))
        return sheetListHang;


    #返回当前sheet的指定列的内容 如果当前列超过了 最大列 或者小于0则会抛出异常
    def getOneColumn(self,whichcolumn):
        if(whichcolumn<0 ):
            raise ColumnIsNotVaild("列 不能小于0");
        if( whichcolumn >self.sheet.get_highest_column()):
            raise ColumnIsNotVaild(str(whichcolumn)+"列已经大于"+str(self.sheet.get_highest_column)+"这是最大列");
        rowLength = self.sheet.get_highest_row();
        sheetListHang  = [];
        # print("colLength"+str(rowLength));
        for j in range(1,rowLength+1):
            sheetListHang.append(self.getCellValueByColumnAndRow(j,whichcolumn));
        return sheetListHang;

    #得到当前sheet表的列的个数
    def getNumbersOfColumn(self):
        return self.sheet.get_highest_column();

    #得到当前sheet表的行的个数
    def getNumbersOfRow(self):
        return self.sheet.get_highest_Row();


    #得到当前excel的表的个数
    def getNumbersOfSheet(self):
        return len(self.wb.get_sheet_names());

    def getNamesOfSheet(self):
        return list(self.wb.get_sheet_names());

    #内部调用  用来转换(B1:C2)->(2,1,3,2)  将所有的合并单元格转换成这种形式 最后加入mergeCellsColumnAndRow列表中
    def __covertStringToColumnAndRow(self):
        # print(self.sheet.merged_cell_ranges)
        self.mergeCellsColumnAndRow = [utils.range_boundaries(x) for x in self.sheet.merged_cell_ranges];
        # print(self.mergeCellsColumnAndRow);


    #内部调用  用来转换(B1:C2)->('B',1,'C',2)  将所有的合并单元格转换成这种形式 最后加入mergeCellsCharAndIndex列表中
    def __setMergeCellsCharAndIndex(self):
        beifen = [utils.range_boundaries(x) for x in self.sheet.merged_cell_ranges];
        for i in beifen:
            t0 = utils.get_column_letter(i[0]);
            t2 = utils.get_column_letter(i[2]);
            self.mergeCellsCharAndIndex.append((t0,i[1],t2,i[3]));
        # print(self.mergeCellsCharAndIndex);

if(__name__=="__main__"):
    filename = r'd:\ceshi.xlsx';
    # read_excel2007(filename);
    a = ReadExcel2007(filename);
    print(a.getCellValueByColumnAndRow(1,2));
    print(a.getCellValueByColumnAndRow(8,6));
    print(a.getCellValueByCharAndRow('G',10));

    a.setCellValueByCharAndRow("G",10,"张");
    a.setCellValueByCharAndRow("AB",10,"张");
    a.setCellValueByColumnAndRow(8,6,"chen");

    a.setCellColorByCharAndRow('G',10,255,0,0);
    a.setCellColorByColumnAndRow(9,6,0,0,255);

    print("&&&&&&&&&&&&&&&&&");
    print(Color(rgb=colors.BLUE).rgb);
    print(a.getCellColorByCharAndRow("f",5));
    print(a.getCellColorByColumnAndRow(5,6));




    # print(a.getAllSheet());
    # print(a.getOneRow(1));
    # print(a.getNamesOfSheet());
    # print(a.getNumbersOfSheet());
    # print(a.getOneColumn(1));
